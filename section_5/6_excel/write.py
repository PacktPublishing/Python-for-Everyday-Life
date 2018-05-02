# -*- coding: utf-8 -*-
# !/usr/bin/env python3

import pathlib
import openpyxl

if __name__ == '__main__':

    composers = [
        {'surname': 'liszt', 'country': 'HU', 'name': 'franz', 'email': 'franz.liszt@gmail.com', 'born': 1811},
        {'surname': 'bach', 'country': 'DE', 'name': 'johann sebastian', 'email': 'jsbach@yahoo.de', 'born': 1685},
        {'surname': 'verdi', 'country': 'IT', 'name': 'giuseppe', 'email': 'giuseppeverdi@hotmail.com', 'born': 1813},
        {'surname': 'gilmour', 'country': 'GB', 'name': 'david', 'email': 'dgilmour@me.com', 'born': 1946}
    ]

    target = pathlib.Path('output.xlsx')
    target.touch()

    # Create workbook and two sheets
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "composers"
    workbook.create_sheet(title='birth')

    # Let's write composers data on the first sheet
    sheet = workbook.get_sheet_by_name('composers')
    sheet.cell(row=1, column=1).value = 'name'
    sheet.cell(row=1, column=2).value = 'surname'
    sheet.cell(row=1, column=3).value = 'country'
    sheet.cell(row=1, column=4).value = 'email'
    for row_number, composer in enumerate(composers):
        sheet.cell(row=row_number+2, column=1).value = composer['name']
        sheet.cell(row=row_number+2, column=2).value = composer['surname']
        sheet.cell(row=row_number+2, column=3).value = composer['country']
        sheet.cell(row=row_number+2, column=4).value = composer['email']


    # Then write birth year on the second sheet
    sheet = workbook.get_sheet_by_name('birth')
    sheet.cell(row=1, column=1).value = 'name'
    sheet.cell(row=1, column=2).value = 'surname'
    sheet.cell(row=1, column=3).value = 'born'
    for row_number, composer in enumerate(composers):
        sheet.cell(row=row_number+2, column=1).value = composer['name']
        sheet.cell(row=row_number+2, column=2).value = composer['surname']
        sheet.cell(row=row_number+2, column=3).value = composer['born']

    # Add a formula to calculate the average of values on column 3 (named "C")
    sheet['B6'].value = 'AVERAGE BIRTH DATE:'
    sheet['C6'] = '=AVERAGE(C2:C5)'

    # save workbook to file
    workbook.save(filename=str(target.resolve()))