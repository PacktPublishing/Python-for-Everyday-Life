# -*- coding: utf-8 -*-
# !/usr/bin/env python3

import pathlib
import openpyxl
import pprint


if __name__ == '__main__':
    src = pathlib.Path('composers.xlsx').resolve()

    print('Loading workbook...')
    workbook = openpyxl.load_workbook(str(src))

    # Sheets in the workbook
    sheets = workbook.get_sheet_names()
    print('\nThe workbook contains the following sheets: {}'.format(sheets))

    # Rows and columns
    composers_sheet = workbook.get_sheet_by_name('composers')
    print('Composers sheet has {} rows and {} columns'.format(
            composers_sheet.max_row,
            composers_sheet.max_column))

    # Let's read composers data into dictionaries
    print('\nExtracting composers data into dictionaries...')
    rows = [r for r in composers_sheet.rows] # .rows gives a generator
    dict_keys = [cell.value for cell in rows[0]]  # header row
    composers = []
    for row in rows[1:]:
        values = [cell.value for cell in row]
        composer = dict(zip(dict_keys, values))
        composers.append(composer)
    pprint.pprint(composers)